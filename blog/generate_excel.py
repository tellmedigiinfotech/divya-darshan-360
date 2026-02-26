import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Base directory
base_dir = r"d:\pratik\newtellme_mediahub\blog"

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Blog Files"

# Define styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Headers
headers = ["Sr. No.", "Folder Name", "File Name", "image1_url", "image2_url"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Set column widths
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 40

# Populate data
row_num = 2
sr_no = 1

for folder_name in sorted(os.listdir(base_dir)):
    folder_path = os.path.join(base_dir, folder_name)
    if os.path.isdir(folder_path):
        files = sorted([f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))])
        for file_name in files:
            ws.cell(row=row_num, column=1, value=sr_no).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_num, column=1).border = thin_border
            
            ws.cell(row=row_num, column=2, value=folder_name).alignment = cell_alignment
            ws.cell(row=row_num, column=2).border = thin_border
            
            ws.cell(row=row_num, column=3, value=file_name).alignment = cell_alignment
            ws.cell(row=row_num, column=3).border = thin_border
            
            ws.cell(row=row_num, column=4, value="").alignment = cell_alignment
            ws.cell(row=row_num, column=4).border = thin_border
            
            ws.cell(row=row_num, column=5, value="").alignment = cell_alignment
            ws.cell(row=row_num, column=5).border = thin_border
            
            row_num += 1
            sr_no += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{row_num - 1}"

# Save
output_path = os.path.join(base_dir, "Blog_Files_List.xlsx")
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")
print(f"Total files listed: {sr_no - 1}")
