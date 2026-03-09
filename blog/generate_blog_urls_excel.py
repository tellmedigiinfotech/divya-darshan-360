#!/usr/bin/env python3
"""
Generate an Excel file listing all blog URLs with a Remarks column.
"""

import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Force UTF-8 output on Windows
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Base directory
base_dir = os.path.dirname(os.path.abspath(__file__))

# Site base URL
BASE_URL = "https://divyadarshan360.com"

# Category mapping (folder name -> URL slug)
CATEGORY_MAP = {
    "Astavinayaka Temple": "astavinayaka",
    "Jyothirlinga Temples": "jyothirlinga",
    "Popular Famous Temple": "popular",
    "Shaktipeet Temple(51)": "shaktipeet",
}

# Category display names
CATEGORY_DISPLAY = {
    "Astavinayaka Temple": "Astavinayaka Temples",
    "Jyothirlinga Temples": "Jyothirlinga Temples",
    "Popular Famous Temple": "Popular Famous Temples",
    "Shaktipeet Temple(51)": "Shaktipeet Temples",
}


def slugify(text):
    """Convert text to URL-friendly slug (matches blog-server.ts logic)."""
    text = text.lower()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'\s+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-').strip()


def get_temple_name(filepath):
    """Extract temple name from the first line of the txt file."""
    try:
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            for line in f:
                stripped = line.strip()
                if stripped:
                    return stripped
    except Exception:
        pass
    return os.path.splitext(os.path.basename(filepath))[0]


# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Blog URLs"

# Define styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="C2560A", end_color="C2560A", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="center", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center")
url_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
category_fill = PatternFill(start_color="FFF3E6", end_color="FFF3E6", fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=11, color="C2560A")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Headers
headers = ["Sr. No.", "Category", "Temple Name", "Blog URL", "Remarks"]
col_widths = [10, 25, 45, 70, 35]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Set column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 35

# Populate data
row_num = 2
sr_no = 1

categories = [
    "Astavinayaka Temple",
    "Jyothirlinga Temples",
    "Popular Famous Temple",
    "Shaktipeet Temple(51)",
]

for cat_folder in categories:
    cat_dir = os.path.join(base_dir, cat_folder)
    if not os.path.isdir(cat_dir):
        print(f"[WARNING] Directory not found: {cat_folder}")
        continue

    category_slug = CATEGORY_MAP.get(cat_folder, slugify(cat_folder))
    category_display = CATEGORY_DISPLAY.get(cat_folder, cat_folder)

    txt_files = sorted([f for f in os.listdir(cat_dir) if f.endswith('.txt')])
    print(f"\n[CATEGORY] {category_display} ({len(txt_files)} blogs)")

    for txt_file in txt_files:
        filepath = os.path.join(cat_dir, txt_file)

        # Get temple name from file content
        temple_name = get_temple_name(filepath)

        # Generate slug from filename (matches blog-server.ts logic)
        filename = txt_file.replace('.docx.txt', '').replace('.txt', '').strip()
        slug = slugify(filename)

        # Build full URL
        blog_url = f"{BASE_URL}/blog/{category_slug}/{slug}"

        # Sr. No.
        cell = ws.cell(row=row_num, column=1, value=sr_no)
        cell.alignment = center_alignment
        cell.border = thin_border

        # Category
        cell = ws.cell(row=row_num, column=2, value=category_display)
        cell.alignment = cell_alignment
        cell.border = thin_border
        cell.font = category_font
        cell.fill = category_fill

        # Temple Name
        cell = ws.cell(row=row_num, column=3, value=temple_name)
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Blog URL (as clickable hyperlink)
        cell = ws.cell(row=row_num, column=4, value=blog_url)
        cell.hyperlink = blog_url
        cell.font = url_font
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Remarks (empty for user to fill)
        cell = ws.cell(row=row_num, column=5, value="")
        cell.alignment = cell_alignment
        cell.border = thin_border

        print(f"  [OK] {temple_name} -> {blog_url}")

        row_num += 1
        sr_no += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{row_num - 1}"

# Save
output_path = os.path.join(base_dir, "Blog_URLs_List.xlsx")
wb.save(output_path)
print(f"\n{'='*60}")
print(f"[DONE] Excel file created: {output_path}")
print(f"[INFO] Total blogs listed: {sr_no - 1}")
