"""Export an Excel sheet of every blog with its public URL and any embedded YouTube link.

Blog name + URL come from public/blog/blogs.json.
YouTube link is extracted from each blog's index.html (iframe src or anchor href).
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]  # frontend/
BLOGS_JSON = ROOT / "public" / "blog" / "blogs.json"
TEMPLE_DIR = ROOT / "public" / "blog" / "temple"
OUTPUT = ROOT / "scripts" / "blogs_youtube.xlsx"

SITE_URL = "https://divyadarshan360.com"

# Mirrors categoryFolderMap in lib/blog-types.ts
CATEGORY_FOLDER_MAP = {
    "popular": "popular",
    "jyothirlinga": "jyothirlinga",
    "shaktipeet": "shaktipeet",
    "Ashtavinayaka": "ashtavinayak",
}

# Match youtube.com/embed/ID, youtube.com/watch?v=ID, youtu.be/ID
YT_PATTERNS = [
    re.compile(r"https?://(?:www\.)?youtube\.com/embed/[A-Za-z0-9_\-]+(?:\?[^\s\"'<>]*)?", re.I),
    re.compile(r"https?://(?:www\.)?youtube\.com/watch\?[^\s\"'<>]*v=[A-Za-z0-9_\-]+[^\s\"'<>]*", re.I),
    re.compile(r"https?://youtu\.be/[A-Za-z0-9_\-]+(?:\?[^\s\"'<>]*)?", re.I),
]


def find_youtube_links(html: str) -> list[str]:
    found: list[str] = []
    for pat in YT_PATTERNS:
        for m in pat.findall(html):
            if m not in found:
                found.append(m)
    return found


def main() -> None:
    blogs = json.loads(BLOGS_JSON.read_text(encoding="utf-8"))

    rows = []
    missing_html = []
    for blog in blogs:
        slug = blog["slug"]
        title = blog["title"].lstrip("﻿").strip()
        category = blog["category"]
        folder = CATEGORY_FOLDER_MAP.get(category, category)
        blog_url = f"{SITE_URL}/blog/temple/{folder}/{slug}"

        index_html = TEMPLE_DIR / folder / slug / "index.html"
        yt_links: list[str] = []
        if index_html.exists():
            try:
                html = index_html.read_text(encoding="utf-8", errors="ignore")
                yt_links = find_youtube_links(html)
            except Exception as exc:  # noqa: BLE001
                yt_links = [f"<read error: {exc}>"]
        else:
            missing_html.append(str(index_html))

        rows.append(
            {
                "Blog Name": title,
                "Category": category,
                "Blog URL": blog_url,
                "Existing YouTube Link(s)": "\n".join(yt_links),
                "Has YouTube": "Yes" if yt_links else "No",
                "New YouTube Link": "",
            }
        )

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Blogs"

    headers = list(rows[0].keys())
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for row in rows:
        ws.append([row[h] for h in headers])

    # Wrap + column widths
    widths = {
        "Blog Name": 55,
        "Category": 14,
        "Blog URL": 75,
        "Existing YouTube Link(s)": 60,
        "Has YouTube": 12,
        "New YouTube Link": 55,
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 25)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    with_yt = sum(1 for r in rows if r["Has YouTube"] == "Yes")
    print(f"Wrote {OUTPUT}")
    print(f"  Total blogs       : {len(rows)}")
    print(f"  With YouTube link : {with_yt}")
    print(f"  Without           : {len(rows) - with_yt}")
    if missing_html:
        print(f"  Missing index.html: {len(missing_html)}")
        for p in missing_html[:10]:
            print("   -", p)


if __name__ == "__main__":
    main()
