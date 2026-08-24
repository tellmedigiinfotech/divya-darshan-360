"""Inject YouTube iframes into temple blog index.html files.

Source: scripts/blogs_youtube.xlsx, column D ("Existing YouTube Link(s)").
Behavior: skip blogs that already have a youtube iframe on disk.
Insertion point: just before `<article class="glass article">`.
"""
from __future__ import annotations

import html as html_mod
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]  # frontend/
SHEET = ROOT / "scripts" / "blogs_youtube.xlsx"
TEMPLE_DIR = ROOT / "public" / "blog" / "temple"

YT_IFRAME_PRESENT = re.compile(
    r'<iframe[^>]*src=["\']https?://(?:www\.)?(?:youtube\.com/embed/|youtu\.be/)',
    re.I,
)

ARTICLE_OPEN = '<article class="glass article">'


def make_block(yt_url: str, title: str) -> str:
    safe_title = html_mod.escape(title.strip(), quote=True)
    return (
        "\t\t\t<!-- Temple Video -->\n"
        '\t\t\t<div class="rounded-xl overflow-hidden shadow-lg border-2 border-[#fdf1e2] mb-12 mt-4 h-[480px]">\n'
        '\t\t\t\t<iframe allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture; web-share"'
        ' allowfullscreen="" class="w-full h-full object-cover" frameborder="0" height="100%"'
        ' referrerpolicy="strict-origin-when-cross-origin"'
        f' src="{yt_url}"'
        f' title="{safe_title} Video"'
        ' width="100%"></iframe>\n'
        "\t\t\t</div>\n\n\t\t\t"
    )


def parse_url(blog_url: str) -> tuple[str, str]:
    """Return (category_folder, slug) from a /blog/temple/<cat>/<slug> URL."""
    parts = blog_url.rstrip("/").split("/")
    return parts[-2], parts[-1]


def main() -> None:
    wb = load_workbook(SHEET, data_only=True)
    ws = wb.active

    targets = []
    for r in range(2, ws.max_row + 1):
        yt = ws.cell(row=r, column=4).value
        if not (yt and isinstance(yt, str) and "youtu" in yt.lower()):
            continue
        title = (ws.cell(row=r, column=1).value or "").lstrip("﻿").strip()
        blog_url = ws.cell(row=r, column=3).value or ""
        cat, slug = parse_url(blog_url)
        targets.append({"row": r, "title": title, "cat": cat, "slug": slug, "yt": yt.strip()})

    print(f"Found {len(targets)} rows with a YouTube link in column D.\n")

    injected, skipped_existing, missing, no_anchor = [], [], [], []

    for t in targets:
        path = TEMPLE_DIR / t["cat"] / t["slug"] / "index.html"
        tag = f'[{t["cat"]}/{t["slug"]}]'

        if not path.exists():
            missing.append(t)
            print(f"  MISSING  {tag}  (file not found: {path.relative_to(ROOT)})")
            continue

        html = path.read_text(encoding="utf-8")

        if YT_IFRAME_PRESENT.search(html):
            skipped_existing.append(t)
            print(f"  SKIP     {tag}  (already has a YouTube iframe)")
            continue

        if ARTICLE_OPEN not in html:
            no_anchor.append(t)
            print(f"  NO-ANCH  {tag}  (could not find article anchor)")
            continue

        block = make_block(t["yt"], t["title"])
        # Insert before the FIRST article open tag
        new_html = html.replace(ARTICLE_OPEN, block + ARTICLE_OPEN, 1)
        path.write_text(new_html, encoding="utf-8")
        injected.append(t)
        print(f"  OK       {tag}  -> {t['yt']}")

    print()
    print("Summary:")
    print(f"  Injected         : {len(injected)}")
    print(f"  Skipped (had iframe): {len(skipped_existing)}")
    print(f"  Missing file     : {len(missing)}")
    print(f"  No article anchor: {len(no_anchor)}")
    if missing:
        print("  -- missing rows:")
        for t in missing:
            print(f"     row {t['row']}: {t['cat']}/{t['slug']}")
    if no_anchor:
        print("  -- no-anchor rows:")
        for t in no_anchor:
            print(f"     row {t['row']}: {t['cat']}/{t['slug']}")


if __name__ == "__main__":
    main()
