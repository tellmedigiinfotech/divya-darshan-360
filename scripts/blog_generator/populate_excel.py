import json
import os
import openpyxl
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "../../"))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "blog_data.xlsx")
JSON_PATH = os.path.join(PROJECT_ROOT, "public", "blog", "blogs.json")

def populate_excel():
    if not os.path.exists(EXCEL_PATH):
        print("Excel not found.")
        return
    if not os.path.exists(JSON_PATH):
        print("JSON not found.")
        return

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        blogs = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["blogs"]
    
    # Read headers
    headers = [cell.value.split('\n')[0] if cell.value else f"col_{i}" for i, cell in enumerate(ws[2])]
    
    # Get existing slugs to avoid duplicates
    existing_slugs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[2]: # slug is index 2
            existing_slugs.append(row[2])

    print(f"Found {len(existing_slugs)} existing entries in Excel.")

    # The template starts with row 4 for new entries
    # But wait, we want to append. Let's find the first empty row
    start_row = 1
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            break
        start_row += 1
    
    current_row = 4
    while any(cell.value for cell in ws[current_row]):
        current_row += 1

    added = 0
    for blog in blogs:
        slug = blog.get("slug")
        if slug in existing_slugs:
            continue
            
        title = blog.get("title", "").replace("\ufeff", "").strip()     
        category = blog.get("category", "")
        location = blog.get("location", "")
        image_url = blog.get("imageUrl", "")
        
        full_image_url = f"https://divyadarshan360.com{image_url}" if image_url.startswith("/") else image_url
        canonical_url = f"https://divyadarshan360.com/blog/temple/{category}/{slug}/index.html"
        
        # Determine category display
        cat_lower = category.lower()
        if cat_lower == "ashtwinayak" or cat_lower == "ashtavinayaka":
            cat_display = "Ashtavinayaka Temples"
            cat_val = "ashtwinayak"
        elif cat_lower == "jyothirlinga":
            cat_display = "Jyotirlinga Temples"
            cat_val = "jyothirlinga"
        elif cat_lower == "shaktipeet":
            cat_display = "Shaktipeeth Temples"
            cat_val = "shaktipeet"
        else:
            cat_display = "Famous Temples"
            cat_val = "popular"
            
        # Parse title to split it
        # Often title is "Something Temple"
        title_main = title
        title_hl = ""
        title_suf = ""
        parts = title.split(" ")
        if len(parts) > 1:
            title_main = " ".join(parts[:-1])
            title_hl = parts[-1]
            
        row_data = {}
        row_data["temple_name"] = title
        row_data["category"] = cat_val
        row_data["slug"] = slug
        row_data["meta_description"] = f"Complete guide to {title} — timings, how to reach, and darshan details."
        row_data["keywords"] = f"{title}, {cat_display}, temple guide, India temples"
        row_data["date_published"] = "2025-01-01T00:00:00+05:30"
        row_data["date_modified"] = "2026-03-31T00:00:00+05:30"
        row_data["og_image_url"] = full_image_url
        row_data["canonical_url"] = canonical_url
        row_data["category_display"] = cat_display
        row_data["hero_title_main"] = title_main
        row_data["hero_title_highlight"] = title_hl
        row_data["hero_image_url"] = full_image_url
        row_data["hero_image_alt"] = f"{title} - View"
        row_data["stat_deity"] = title
        row_data["stat_location"] = location
        row_data["main_deity"] = title
        
        # Write to excel
        for col_idx, key in enumerate(headers):
            if key in row_data:
                ws.cell(row=current_row, column=col_idx+1, value=row_data[key])
                
        current_row += 1
        added += 1
        
    wb.save(EXCEL_PATH)
    print(f"Added {added} new blogs to Excel.")

if __name__ == '__main__':
    populate_excel()
