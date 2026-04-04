"""
analyze_missing.py
Analyzes blog_data.xlsx for missing fields per temple row.
"""
import sys
import io
import os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "blog_data.xlsx")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['blogs']

print(f"Sheet: blogs | Rows: {ws.max_row} | Cols: {ws.max_column}\n")

# Extract headers from row 2 (key name is before the newline)
headers = []
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=2, column=c).value
    if v:
        key = str(v).split("\n")[0].strip()
    else:
        key = f"col{c}"
    headers.append(key)

print("=== COLUMNS ===")
for i, h in enumerate(headers):
    print(f"  Col {i+1}: {h}")

print("\n=== ROWS WITH DATA ===")
for r in range(3, ws.max_row + 1):
    temple = ws.cell(row=r, column=1).value
    if temple and str(temple).strip():
        missing = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip() == "":
                missing.append(headers[c - 1])
        print(f"\nRow {r}: {temple}")
        if missing:
            print(f"  MISSING ({len(missing)}): {missing}")
        else:
            print(f"  All fields complete!")
    else:
        print(f"Row {r}: [EMPTY]")
