import os
import openpyxl
from jinja2 import Template
import sys

# Encoding fix for Windows
sys.stdout.reconfigure(encoding='utf-8')

# Configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "blog_data.xlsx")
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "blog_template.html.jinja")
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "../../"))

def generate_blogs():
    print(f"[PROCESS] Starting blog generation...")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel file not found: {EXCEL_PATH}")
        return

    if not os.path.exists(TEMPLATE_PATH):
        print(f"[ERROR] Template file not found: {TEMPLATE_PATH}")
        return

    # Load Template
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template_str = f.read()
    template = Template(template_str)

    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["blogs"]
    
    # Get headers from Row 2
    headers = [cell.value.split('\n')[0] if cell.value else f"col_{i}" for i, cell in enumerate(ws[2])]
    
    gen_count = 0
    err_count = 0
    
    # Process rows starting from Row 4 (skipping headers and example row 3)
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Create data dictionary for this row
        data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                key = headers[i]
                # Default empty values to empty string
                data[key] = str(value).strip() if value is not None else ""

        # Validate required fields
        if not data.get("temple_name") or not data.get("slug") or not data.get("category"):
            # Check if row is just empty
            if all(not val for val in data.values()):
                continue
            print(f"[SKIP] Row {row_idx}: Missing temple_name, category, or slug.")
            err_count += 1
            continue

        try:
            # Render HTML
            html_output = template.render(**data)
            
            # Create output path: public/blog/temple/[category]/[slug]/index.html
            output_dir = os.path.join(PROJECT_ROOT, "public", "blog", "temple", data["category"], data["slug"])
            os.makedirs(output_dir, exist_ok=True)
            
            output_file = os.path.join(output_dir, "index.html")
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(html_output)
            
            print(f"[{gen_count+1}] SUCCESS: public/blog/temple/{data['category']}/{data['slug']}/index.html")
            gen_count += 1
            
        except Exception as e:
            print(f"[ERROR] Row {row_idx} ({data.get('temple_name')}): {str(e)}")
            err_count += 1

    print(f"\n[FINISH] Generation completed.")
    print(f"   - Successfully generated: {gen_count}")
    print(f"   - Failed/Skipped: {err_count}")
    print(f"\nInstructions: View the generated blogs in the browser to verify design.")

if __name__ == "__main__":
    generate_blogs()
