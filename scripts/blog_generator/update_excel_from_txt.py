import os
import openpyxl
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "blog_data.xlsx")
TXT_BLOGS_DIR = os.path.join(SCRIPT_DIR, "txt_blogs")

def clean_val(text):
    if not text: return ""
    text = text.replace("\ufeff", "").replace("•", "").strip()
    text = re.sub(r'\n\s*\n', '\n', text)
    return text

def parse_timings(text, updates):
    if not text: return
    # Fill full timing
    updates["temple_full_timing"] = text
    # Try extract specific AM/PM blocks
    am_pm = re.findall(r'\d{1,2}:\d{2}\s*[A-Z]{2}', text)
    if len(am_pm) >= 2:
        updates["timing_morning"] = f"{am_pm[0]} - {am_pm[1]}"
        if len(am_pm) >= 4:
            updates["timing_evening"] = f"{am_pm[2]} - {am_pm[3]}"
        else:
            updates["timing_evening"] = f"5:00 PM - 9:00 PM"
    else:
        # Defaults for basic temples
        updates["timing_morning"] = "6:00 AM - 12:00 PM"
        updates["timing_evening"] = "5:00 PM - 9:00 PM"

def parse_poojas(text, updates):
    if not text: return
    lines = text.split('\n')
    idx = 1
    for line in lines:
        if ":" in line and idx <= 3:
            updates[f"special_puja_{idx}"] = line.strip().replace(":", "|")
            idx += 1

def smart_extract(content, keywords, max_chars=1000):
    if isinstance(keywords, str): keywords = [keywords]
    content_lines = content.split('\n')
    for i, line in enumerate(content_lines):
        for kw in keywords:
            if re.search(rf"{re.escape(kw)}", line, re.IGNORECASE):
                result_lines = []
                if ":" in line:
                    parts = line.split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                        result_lines.append(parts[1].strip())
                for j in range(i + 1, min(i + 20, len(content_lines))):
                    next_line = content_lines[j].strip()
                    if not next_line:
                        if len(result_lines) > 2: break
                        continue
                    if re.match(r'^(?:[^\w\s]{1,4})?\s*[A-Z][A-Za-z ]{2,30}[:\-\?]?$', next_line):
                        break
                    result_lines.append(next_line)
                res = "\n".join(result_lines).strip()
                if res and len(res) > 5:
                    return res[:max_chars]
    return ""

def update_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["blogs"]
    headers = [cell.value.split('\n')[0] if cell.value else f"col_{i}" for i, cell in enumerate(ws[2])]
    header_map = {name: i + 1 for i, name in enumerate(headers)}
    
    processed = 0
    for r in range(4, ws.max_row + 1):
        temple_name = str(ws.cell(row=r, column=header_map.get("temple_name", 1)).value or "")
        slug = str(ws.cell(row=r, column=header_map.get("slug", 3)).value or "")
        if not slug: continue
        txt_file = None
        for root, _, files in os.walk(TXT_BLOGS_DIR):
            for file in files:
                if file.endswith(".txt") and (slug in file or slug[:20] in file):
                    txt_file = os.path.join(root, file)
                    break
            if txt_file: break
        if not txt_file: continue
        processed += 1
        with open(txt_file, "r", encoding="utf-8") as f: content = f.read()
        u = {}
        u["stat_contact"] = smart_extract(content, ["Contact", "Phone"])
        u["address_full"] = smart_extract(content, ["Full Address", "Location"])
        u["main_deity_desc"] = smart_extract(content, "Main Deity")
        u["arch_style_desc"] = smart_extract(content, ["Construction Features", "Architectural Features", "Historical Background"])
        u["arch_scientific_desc"] = smart_extract(content, "Scientific Insights")
        u["significance_para_1"] = smart_extract(content, ["Religious Significance", "Spiritual significance", "Chintamani Legend"])
        u["hero_intro_para_1"] = smart_extract(content, ["The temple", "Situated", "Spiritual Experience", "Overview"])[:500]
        
        # Parse Timings and Poojas
        timings_raw = smart_extract(content, ["Opening & Closing", "Timings", "Regular Darshan"])
        parse_timings(timings_raw, u)
        
        poojas_raw = smart_extract(content, ["Special Poojas", "Daily Poojas", "Regular Monthly Rituals"])
        parse_poojas(poojas_raw, u)
        
        # Foreigner cards cleanup
        env = smart_extract(content, "Spiritual Environment")
        if env: u["foreign_card_1"] = f"Spiritual Environment|{env[:300]}"
        exp = smart_extract(content, "Experience and Practices")
        if exp: u["foreign_card_2"] = f"Experience and Practices|{exp[:300]}"
        phi = smart_extract(content, "Philosophical Insights") or smart_extract(content, "Mythological Context")
        if phi: u["foreign_card_3"] = f"Philosophical Insights|{phi[:300]}"

        for k, v in u.items():
            if k in header_map and v:
                ws.cell(row=r, column=header_map[k], value=clean_val(v))
    wb.save(EXCEL_PATH)
    print(f"Extraction complete for {processed} temples.")

if __name__ == "__main__":
    update_excel()
