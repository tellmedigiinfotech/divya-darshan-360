"""
add_maps_links_report.py
------------------------
Reads Geo Codes from link_update.xlsx (Sheet2) and adds a Google Maps
link to the address/location section of every temple blog index.html.

Generates an Excel report (maps_update_report.xlsx) with:
1. "Have Geo Code" - List of temples where geo code was found and updated (or skipped).
2. "No Geo Code Found" - List of temples with no geo code.
3. "HTML Match Failed" - List of temples where HTML pattern was not found.
"""

import os, glob, re
import openpyxl
from openpyxl import Workbook

EXCEL     = r'd:\pratik\divya_darshan\divya-darshan-360\scripts\blog_generator\link_update.xlsx'
BLOG_BASE = r'd:\pratik\divya_darshan\divya-darshan-360\public\blog\temple'
REPORT_EXCEL = r'd:\pratik\divya_darshan\divya-darshan-360\scripts\blog_generator\maps_update_report.xlsx'

MAPS_BASE = "https://www.google.com/maps/search/"

# ── 1. Load Geo-code map from Sheet2 ─────────────────────────────────────────
def load_geo_map(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws2 = wb.worksheets[1]      # Sheet2 (0-indexed)
    geo_map = {}                # lowercase_name  →  "lat, lng"
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i < 2:               # skip two header rows
            continue
        if row[0] is None:      # stop at first empty Sr.No
            break
        name      = row[2]      # column C — temple name
        geo_codes = row[6]      # column G — geo codes
        if name and geo_codes:
            key = name.lower().strip()
            geo_map[key] = str(geo_codes).strip()
    wb.close()
    return geo_map

# ── 2. Match a slug to a geo code ────────────────────────────────────────────
def find_geo(slug, geo_map):
    for name, geo in geo_map.items():
        if name.replace(" ", "-") in slug or name.replace(" ", "") in slug.replace("-", ""):
            return geo, name
    return None, None

# ── 3. Maps link HTML snippet ─────────────────────────────────────────────────
def maps_link_html(geo_code):
    url = MAPS_BASE + geo_code.replace(" ", "")
    return (
        f'\n<a href="{url}" target="_blank" rel="noopener noreferrer" '
        f'style="display:inline-flex;align-items:center;gap:4px;margin-top:6px;'
        f'font-size:.82rem;color:#e23d00;text-decoration:none;font-weight:500;" '
        f'class="maps-link">'
        f'<svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" '
        f'stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">'
        f'<path d="M21 10c0 7-9 13-9 13s-9-6-9-13a9 9 0 0 1 18 0z"></path>'
        f'<circle cx="12" cy="10" r="3"></circle></svg>'
        f'View on Google Maps</a>'
    )

# ── 4. Patch one HTML file ────────────────────────────────────────────────────
SIDEBAR_LOC_RE = re.compile(
    r'(<p class="text-\[\.75rem\] font-normal text-primary uppercase tracking-widest mb-1">Location</p>\s*'
    r'<p class="text-\[\.88rem\] text-\[#3b2510\]">.*?</p>)',
    re.DOTALL
)

REACH_P_RE = re.compile(
    r'(<p class="text-\[\.9rem\] mb-1">(?:[^<]|<(?!/p>))*?</p>)(\s*<p class="text-\[\.9rem\]"><strong>Phone)',
    re.DOTALL
)

def already_patched(content):
    return 'maps-link' in content or 'maps/search' in content

def patch_file(file_path, geo_code):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    if already_patched(content):
        return 'skipped'

    link_html = maps_link_html(geo_code)
    changed = False

    def replace_sidebar(m):
        nonlocal changed
        changed = True
        return m.group(1) + link_html

    new_content = SIDEBAR_LOC_RE.sub(replace_sidebar, content, count=1)

    def replace_reach(m):
        nonlocal changed
        changed = True
        return m.group(1) + link_html + m.group(2)

    new_content = REACH_P_RE.sub(replace_reach, new_content, count=1)

    if changed:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        return 'updated'
    return 'no_match'

# ── 5. Main ───────────────────────────────────────────────────────────────────
def main():
    geo_map = load_geo_map(EXCEL)

    html_files = glob.glob(os.path.join(BLOG_BASE, '**', 'index.html'), recursive=True)
    
    # Report lists
    have_geo_list = []
    no_geo_list = []
    html_match_failed_list = []
    
    print(f"Processing {len(html_files)} HTML files...")

    for f in sorted(html_files):
        slug = os.path.basename(os.path.dirname(f)).lower()
        geo, matched_name = find_geo(slug, geo_map)

        if not geo:
            no_geo_list.append([slug, ""])
            continue

        result = patch_file(f, geo)
        
        if result in ('updated', 'skipped'):
            have_geo_list.append([slug, matched_name, geo, result])
        elif result == 'no_match':
            html_match_failed_list.append([slug, matched_name, geo])

    # Generate Report Excel
    wb = Workbook()
    
    # Sheet 1: Have Geo Code
    ws1 = wb.active
    ws1.title = "Have Geo Code"
    ws1.append(["Slug", "Matched Name", "Geo Code", "Status"])
    for row in have_geo_list:
        ws1.append(row)
        
    # Sheet 2: No Geo Code Found
    ws2 = wb.create_sheet(title="No Geo Code Found")
    ws2.append(["Slug", "Note"])
    for row in no_geo_list:
        ws2.append(row)
        
    # Sheet 3: HTML Match Failed
    ws3 = wb.create_sheet(title="HTML Match Failed")
    ws3.append(["Slug", "Matched Name", "Geo Code"])
    for row in html_match_failed_list:
        ws3.append(row)
        
    wb.save(REPORT_EXCEL)
    print(f"\nDone! Report saved to: {REPORT_EXCEL}")
    
    print(f"\nSummary:")
    print(f"  Have Geo Code: {len(have_geo_list)}")
    print(f"  No Geo Code: {len(no_geo_list)}")
    print(f"  HTML Match Failed: {len(html_match_failed_list)}")

if __name__ == '__main__':
    main()
