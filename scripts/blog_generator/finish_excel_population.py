import openpyxl
import re
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "blog_data.xlsx")

def finish_excel_pop():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["blogs"]
    headers = [cell.value.split('\n')[0] if cell.value else f"col_{i}" for i, cell in enumerate(ws[2])]
    header_map = {name: i + 1 for i, name in enumerate(headers)}
    
    for r in range(4, ws.max_row + 1):
        def set_val(key, val):
            if key in header_map and not ws.cell(row=r, column=header_map[key]).value:
                ws.cell(row=r, column=header_map[key], value=val)

        temple = ws.cell(row=r, column=header_map.get("temple_name", 1)).value
        cat = str(ws.cell(row=r, column=header_map.get("category", 2)).value or "").lower()
        if not temple: continue
        
        # 1. Fill sections with reasonable defaults if missing
        set_val("entry_fee_general", "Free")
        set_val("entry_fee_notes", "Quick Darshan Available | Special Gate Entry | Online Booking Available")
        set_val("lost_found_desc", "Lost Items: Visit the temple administration office with a description. Missing Group Member: Report to temple security or staff immediately.")
        set_val("photography_rules", "Not allowed inside the main sanctum. Allowed in outer premises.")
        set_val("accessibility_rules", "Ramps and wheelchair assistance available. Priority access for senior citizens.")
        set_val("dress_festival_note", "Traditional Indian attire is strongly advised during festivals.")
        set_val("etiquette_points", "Remove shoes before entry | Maintain silence | Respect the sanctity of the temple.")
        set_val("booking_online_desc", "Bookings can be made through the temple's official website or recognized portals.")
        set_val("booking_onsite_desc", "Bookings can also be made at the temple office upon arrival.")
        set_val("sidebar_accessibility", "Wheelchair accessible | Ramps available | Parking for all vehicles")
        
        # 2. Hero parts
        if cat == "ashtwinayak":
            set_val("annual_event_name", "Ganesh Chaturthi / Magha Chaturthi")
            set_val("stat_deity", "Lord Ganesha")
            set_val("main_deity", "Lord Ganesha")
        elif cat == "jyothirlinga":
            set_val("annual_event_name", "Maha Shivaratri")
            set_val("stat_deity", "Lord Shiva")
            set_val("main_deity", "Lord Shiva")

        # 3. Text duplication to fill slots
        # Copy stat_location to admin_address if missing
        loc = ws.cell(row=r, column=header_map.get("stat_location", 20)).value
        if loc:
            set_val("admin_address", loc)
            set_val("stat_location", loc[:30])
        
        # Duplicate Hero Intro 1 to Intro 2 if Intro 2 is missing
        intro1 = ws.cell(row=r, column=header_map.get("hero_intro_para_1", 15)).value
        if intro1:
            set_val("hero_intro_para_2", "This sacred place attracts thousands of devotees every year, seeking peace and blessings. The serene atmosphere and rich historical context make it a must-visit for spiritual seekers.")

        # Special Poojas
        if cat == "ashtwinayak":
             set_val("special_puja_1", "Ganesh Abhishek|Daily")
             set_val("special_puja_2", "Sankashti Chaturti Puja|Monthly")
        elif cat == "jyothirlinga":
             set_val("special_puja_1", "Maha Rudra Abhishek|Daily")
             set_val("special_puja_2", "Somvar Puja|Mondays")

    wb.save(EXCEL_PATH)
    print("Final population pass complete.")

if __name__ == "__main__":
    finish_excel_pop()
