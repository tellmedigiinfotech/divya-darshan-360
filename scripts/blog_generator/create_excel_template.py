"""
create_excel_template.py
========================
Run this once to create/reset the Excel template file.
It creates blog_data.xlsx with Ballaleshwar as the example row.

Usage:
    python scripts/blog_generator/create_excel_template.py
"""

import os
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Output path
# ─────────────────────────────────────────────
OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "blog_data.xlsx"
)

# ─────────────────────────────────────────────
# Column definitions: (header_key, display_label, section_color, example_value)
# ─────────────────────────────────────────────
COLUMNS = [
    # ── SECTION A: Basic Info & SEO ─────────────────────────────────────────
    ("temple_name",       "Temple Name",                  "FF1E3A5F", "Ashtavinayaka Ballaleshwar Temple"),
    ("category",          "Category (URL folder)",        "FF1E3A5F", "ashtwinayak"),
    ("slug",              "Slug (folder name)",           "FF1E3A5F", "ballaleshwar-temple-found-in-pali-village-raigad-district"),
    ("meta_description",  "Meta Description (~155 chars)","FF1E3A5F", "Complete guide to Ashtavinayaka Ballaleshwar Temple — timings, how to reach, darshan booking, poojas, dress code, and facilities. Located in Maharashtra -410205, India. Plan your spiritual visit today."),
    ("keywords",          "Keywords (comma-separated)",   "FF1E3A5F", "Ashtavinayaka Ballaleshwar Temple, Ashtavinayaka Ballaleshwar Temple timings, Ganesh temples Maharashtra, Ashtavinayak darshan"),
    ("date_published",    "Date Published (ISO)",         "FF1E3A5F", "2025-01-01T00:00:00+05:30"),
    ("date_modified",     "Date Modified (ISO)",          "FF1E3A5F", "2026-03-24T00:00:00+05:30"),
    ("og_image_url",      "OG Image URL (full URL)",      "FF1E3A5F", "https://divyadarshan360.com/blog/images/Ashtavinayaka Temple/ballaleshwar-temple-found-in-pali-village-raigad-district_1.jpg"),
    ("canonical_url",     "Canonical URL (full URL)",     "FF1E3A5F", "https://divyadarshan360.com/blog/temple/ashtwinayak/ballaleshwar-temple-found-in-pali-village-raigad-district/index.html"),
    ("category_display",  "Category Badge Text",          "FF1E3A5F", "Ashtavinayaka Temples"),

    # ── SECTION B: Hero ─────────────────────────────────────────────────────
    ("hero_title_main",      "H1 Title Part 1 (plain)",       "FF2D6A4F", "Ashtavinayaka"),
    ("hero_title_highlight", "H1 Title Part 2 (colored)",     "FF2D6A4F", "Ballaleshwar"),
    ("hero_title_suffix",    "H1 Title Part 3 (after color)", "FF2D6A4F", "Temple"),
    ("hero_intro_para_1",    "Hero Intro Paragraph 1",        "FF2D6A4F", "The Ballaleshwar Temple in Pali village, Raigad District, Maharashtra is one of the eight sacred Ashtavinayak temples of Lord Ganesha. It holds immense historical and religious significance, drawing thousands of devotees and pilgrims every year seeking blessings and spiritual enlightenment."),
    ("hero_intro_para_2",    "Hero Intro Paragraph 2",        "FF2D6A4F", "Standing as a testament to the rich architectural heritage and religious traditions of Maharashtra, the temple attracts visitors and spiritual seekers from far and wide. Many come to reflect, connect with the temple's serene environment, and experience its powerful spiritual aura."),
    ("hero_image_url",       "Hero Image URL",                "FF2D6A4F", "https://divyadarshan360.com/blog/images/Ashtavinayaka Temple/ballaleshwar-temple-found-in-pali-village-raigad-district_1.jpg"),
    ("hero_image_alt",       "Hero Image Alt Text",           "FF2D6A4F", "Ashtavinayaka Ballaleshwar Temple - Front View"),

    # ── SECTION C: Quick Stat Cards ─────────────────────────────────────────
    ("stat_deity",    "Main Deity",          "FF6D4C41", "Lord Ganesha (Ballaleshwar)"),
    ("stat_location", "Short Location",      "FF6D4C41", "Pali Village, Sudhagad, Raigad, Maharashtra"),
    ("stat_contact",  "Contact / Phone",     "FF6D4C41", "02142-242263"),

    # ── SECTION D: Overview Tab ──────────────────────────────────────────────
    ("main_deity",        "Main Deity (Short)",          "FF4A235A", "Lord Ganesha (Ballaleshwar)"),
    ("main_deity_desc",   "Main Deity Detail",           "FF4A235A", "Lord Ganesha is worshipped here as Ballaleshwar, uniquely dressed as a Brahmin with a left-turning trunk."),
    ("sub_deities",       "Sub Deities (pipe | separated)",              "FF4A235A", "Goddess Parvati|Lord Shiva"),
    ("sub_deities_desc",  "Sub Deities Detail",          "FF4A235A", "The temple compound houses several smaller shrines dedicated to various associated deities."),
    ("temple_trust",      "Temple Trust Name",                           "FF4A235A", "Ballaleshwar Temple Trust"),
    ("temple_trust_desc", "Temple Trust Detail",         "FF4A235A", "Managed by the official local devasthan committee responsible for daily pujas and festivals."),
    ("historical_para_1", "Historical & Arch. Para 1",                   "FF4A235A", "The Ballaleshwar Temple is constructed in the traditional Hindu architectural style prevalent in Maharashtra. It features a shikara (tower) atop the main sanctum (garbhagriha), surrounded by various mandapas (halls) and corridors. The temple is primarily built using stone and mortar, with intricate carvings adorning its walls and pillars."),
    ("historical_para_2", "Historical & Arch. Para 2",                   "FF4A235A", "The temple is associated with the legend of Ballala, a devoted worshipper of Lord Ganesha whose unwavering faith is said to have pleased the deity, leading to this sacred shrine being established in his honor. Conservation projects for structural repairs, restoration of carvings, and preservation of historical artifacts ensure its heritage is maintained."),
    ("spiritual_para_1",  "Spiritual Significance Para 1",               "FF4A235A", "Lord Ganesha is revered as the remover of obstacles and the patron of intellect and wisdom. The temple's natural surroundings — lush greenery, serene atmosphere, and nearby hills — evoke a sense of tranquility and oneness with nature. Devotees and spiritual seekers alike visit to participate in meditation, chanting of mantras, and bhajans (devotional songs)."),
    ("spiritual_para_2",  "Spiritual Significance Para 2",               "FF4A235A", "For foreign visitors, this temple offers a journey of spiritual exploration. Engaging with the symbolism of Lord Ganesha, observing spiritual practices, and connecting with the temple's peaceful environment can foster personal growth transcending religious and cultural boundaries."),
    ("entry_fee_general", "General Entry Fee",                           "FF4A235A", "Free"),
    ("entry_fee_notes",   "Fee Rows (Label|Value|Label|Value...)",        "FF4A235A", "Quick Darshan|Available for specific rituals or donations|Special Gate Entry|Designated gate for quick entry|Extended Deity Viewing|During special poojas|Prasad|Special prasad for certain donations|Puja Participation|Opportunity to join specific rituals"),
    ("annual_event_name", "Annual Event Name",                           "FF4A235A", "Magha Shukla Chaturthi"),
    ("annual_event_desc", "Annual Event Description",                    "FF4A235A", "Celebrated on the fourth day of the waxing moon in the Hindu month of Magha. Specific dates vary annually with the lunar calendar — please verify with the temple directly."),
    ("monthly_event_name","Monthly Event Name",                          "FF4A235A", "Pradosha Puja"),
    ("monthly_event_desc","Monthly Event Description",                   "FF4A235A", "Observed twice monthly — on the 13th day of both the waxing (Shukla Paksha) and waning (Krishna Paksha) lunar phases. These are auspicious days for special prayers and rituals."),
    ("donation_channels", "Donation Channels (BoldLabel:desc|...)",      "FF4A235A", "Online Donations:|Through the temple's official website using credit/debit cards, net banking, or other online payment methods.|Bank Transfers:|Designated bank accounts; details available on the temple's official website.|Cheque / Demand Draft:|Written in favour of the temple or the Ballaleshwar Temple Trust.|Mobile Payment Apps:|The temple may have tie-ups with certain mobile payment platforms."),

    # ── SECTION E: Facilities Tab ───────────────────────────────────────────
    ("facility_1",      "Facility 1 (emoji|Title|Desc)",  "FF1A5276", "♿|Accessibility|Ramps and wheelchair assistance are available throughout the temple premises. Priority access and special timings are provided for sick, senior citizens, and physically challenged individuals."),
    ("facility_2",      "Facility 2 (emoji|Title|Desc)",  "FF1A5276", "🦽|Wheelchair Service|Wheelchairs can be requested at the temple entrance or help desk upon arrival. Temple staff assist with navigation throughout the premises."),
    ("facility_3",      "Facility 3 (emoji|Title|Desc)",  "FF1A5276", "🏠|Accommodation|Guest houses (basic & deluxe rooms) and dormitories are available for pilgrims. Book online via the temple's website or on-site at the accommodation office."),
    ("facility_4",      "Facility 4 (emoji|Title|Desc)",  "FF1A5276", "🚗|Parking|Designated parking areas for cars, specific spots for mini buses / large vehicles, and separate spaces for two-wheelers — all managed for convenient access."),
    ("facility_5",      "Facility 5 (emoji|Title|Desc)",  "FF1A5276", "🍜|Prasad & Meals|Free basic prashad distribution. Special prashad (Ladoos, Modaks) available for purchase. Meal timings align with temple schedule."),
    ("facility_6",      "Facility 6 (emoji|Title|Desc)",  "FF1A5276", "🔒|Gadget Storage|Electronic gadgets (mobile phones, cameras, laptops) are generally not allowed inside the main temple. A designated cloakroom / locker facility is available near the temple entrance."),
    ("online_services", "Online Services (BoldLabel:desc|...)","FF1A5276","Puja Booking:|Book various poojas and rituals online.|Donations:|Secure online donation options available.|Live Darshan:|Streaming services for real-time darshan.|Accommodation Booking:|Book guest houses and other lodgings online."),
    ("vehicle_pooja_desc","Vehicle Pooja Description",     "FF1A5276", "Vehicle Pooja is performed at the temple during operational hours (5:00 AM – 10:00 PM). Contact the temple administration directly for current rates and scheduling."),
    ("lost_found_desc", "Lost & Found Description",        "FF1A5276", "Lost Items: Visit the temple administration office, provide description of lost item, fill out a form, and provide your contact info. Missing Group Member: Report immediately to temple security or staff."),

    # ── SECTION F: Guidelines Tab ────────────────────────────────────────────
    ("dress_men_do",      "Dress Code Men ✓ (pipe-sep)",   "FF117A65", "Traditional attire is recommended."),
    ("dress_men_dont",    "Dress Code Men ✗ (pipe-sep)",   "FF117A65", "Avoid shorts and sleeveless tops."),
    ("dress_women_do",    "Dress Code Women ✓ (pipe-sep)", "FF117A65", "Sarees or salwar kameez are preferred."),
    ("dress_women_dont",  "Dress Code Women ✗ (pipe-sep)", "FF117A65", "Avoid short skirts, sleeveless tops, and tight clothing."),
    ("dress_festival_note","Festival Attire Note",         "FF117A65", "Traditional Indian attire is strongly advised to show respect and enhance the spiritual experience."),
    ("etiquette_points",  "Temple Etiquette ✓ (pipe-sep)", "FF117A65", "Remove shoes before entering the sanctum.|Maintain silence and peaceful conduct throughout the premises.|Offer prayers with sincerity and respect.|Respect the sanctity of the temple during prayers and rituals."),
    ("photography_rules", "Photography Rules (BoldLabel:desc|...)","FF117A65","Photography inside the temple:|Not allowed, to maintain the sanctity of the premises.|Photography outside the temple:|Allowed in the outer premises and surrounding areas. Generally no fee.|Professional Photography / Videography:|Requires prior permission from temple authorities.|Electronic Gadgets (mobiles, cameras, laptops):|Generally not allowed inside the main temple."),
    ("accessibility_rules","Accessibility Rules (pipe-sep)","FF117A65","Ramps and wheelchair assistance are available for physically handicapped visitors.|No lift/elevator is specifically mentioned within the temple premises.|Priority access and special darshan timings for sick persons, senior citizens, and physically challenged individuals.|Wheelchairs can be requested at the temple entrance or help desk."),

    # ── SECTION G: How to Reach ──────────────────────────────────────────────
    ("address_full",     "Full Address",                   "FF784212", "Pali Village, Sudhagad Taluka, Raigad District, Maharashtra – 410205, India"),
    ("address_phone",    "Address Phone",                  "FF784212", "02142-242263"),
    ("air_airport_name", "Nearest Airport Name",           "FF784212", "Navi Mumbai International Airport"),
    ("air_map_embed",    "By Air — Google Maps Embed URL", "FF784212", "https://www.google.com/maps/embed?pb=!1m18!1m12!1m3!1d601115.843073749!2d72.82709414454185!3d18.81401034139885!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!3m3!1m2!1s0x3be7c3007110c111%3A0x901952e74840c074!2sNavi%20Mumbai%20International%20Airport!5e1!3m2!1sen!2sin!4v1774614670076!5m2!1sen!2sin"),
    ("train_station_name","Nearest Train Station & Distance","FF784212","Khopoli Railway Station (~30 km)"),
    ("train_map_embed",  "By Train — Google Maps Embed URL","FF784212","https://www.google.com/maps/embed?pb=!1m18!1m12!1m3!1d4696.9305638103915!2d73.34346207519825!3d18.78845983235789!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!3m3!1m2!1s0x3be8070024b61999%3A0x5e4fc0e63aff9779!2sKhopoli%20railway%20station!5e1!3m2!1sen!2sin!4v1774614226255!5m2!1sen!2sin"),
    ("bus_stand_name",   "Bus Stand Name & Routes",        "FF784212", "Pali Bus Stand — from Mumbai & Pune"),
    ("bus_map_embed",    "By Bus — Google Maps Embed URL", "FF784212", "https://www.google.com/maps/embed?pb=!1m18!1m12!1m3!1d4703.803332366822!2d73.21530462519212!3d18.540446432557033!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!3m3!1m2!1s0x3be8114653f708a1%3A0x4cf9c9f5b824f63b!2sPali%20Bus%20Stand!5e1!3m2!1sen!2sin!4v1774614764536!5m2!1sen!2sin"),
    ("taxi_desc",        "Taxi / Auto Description",        "FF784212", "Taxis and auto-rickshaws are readily available from Khopoli Railway Station and surrounding towns directly to the temple."),

    # ── SECTION H: YouTube ───────────────────────────────────────────────────
    ("youtube_video_id", "YouTube Video ID only",          "FF922B21", "pAtgE0CMKQs"),

    # ── SECTION I: Timings ───────────────────────────────────────────────────
    ("timing_morning",       "Morning Darshan Timing",          "FF0E6655", "5:00 AM – 8:00 AM"),
    ("timing_midday",        "Midday Darshan Timing",           "FF0E6655", "12:00 PM – 2:00 PM"),
    ("timing_evening",       "Evening Darshan Timing",          "FF0E6655", "6:00 PM – 9:00 PM"),
    ("special_puja_1",       "Special Puja 1 (Name|Day)",       "FF0E6655", "Rudra Abhisheka|Every Friday"),
    ("special_puja_2",       "Special Puja 2 (Name|Day)",       "FF0E6655", "Bhajana Seva|Every Saturday"),
    ("special_puja_3",       "Special Puja 3 (Name|Day)",       "FF0E6655", "Eshwar Dwija Pooja|Sundays"),
    ("temple_full_timing",   "Full Temple Timing (Sidebar)",    "FF0E6655", "5:00 AM – 10:00 PM"),
    ("timing_extended_note", "Festival Timing Note",            "FF0E6655", "Extended on festival days"),

    # ── SECTION J: Architecture & Significance ───────────────────────────────
    ("arch_style_desc",       "Architectural Style Para",       "FF1F618D", "The Ballaleshwar Temple is constructed in the traditional Hindu architectural style prevalent in Maharashtra. It features a shikara (tower) atop the main sanctum (garbhagriha), surrounded by various mandapas (halls) and corridors."),
    ("arch_material_desc",    "Building Material Para",         "FF1F618D", "The temple is primarily constructed using stone and mortar, with intricate carvings adorning its walls and pillars. The use of stone provides durability and stability to the structure, ensuring its longevity."),
    ("arch_scientific_desc",  "Scientific Aspects Para",        "FF1F618D", "Specific geometrical proportions and architectural techniques enhance the temple's structural integrity and help withstand environmental factors such as earthquakes and weathering."),
    ("arch_conservation_desc","Conservation Efforts Para",      "FF1F618D", "Given its historical and religious significance, ongoing conservation projects involve structural repairs, restoration of carvings, and preservation of historical artifacts associated with the temple."),
    ("significance_para_1",   "Special Significance Para 1",   "FF1F618D", "The Ballaleshwar Temple's significance lies in its deep historical and religious importance, particularly as a place of worship dedicated to Lord Ganesha and its association with the legend of Ballala — a devotee whose unwavering faith is said to have pleased the deity."),
    ("significance_para_2",   "Special Significance Para 2",   "FF1F618D", "Many temples in India are believed to be located in places with unique spiritual energies or natural phenomena. These beliefs hold deep cultural and religious significance for devotees and locals."),
    ("significance_para_3",   "Special Significance Para 3",   "FF1F618D", "Devotees visit the temple to seek blessings, offer prayers, and participate in festivals and ceremonies held throughout the year, contributing to its enduring spiritual atmosphere and cultural heritage."),

    # ── SECTION K: For Foreign Visitors ─────────────────────────────────────
    ("foreign_card_1", "Foreign Visitor Card 1 (Title|Desc)", "FF4A235A", "Historical Legend|Understanding the legend of Ballala and his unwavering devotion offers insight into the spiritual significance of the temple beyond its religious practices."),
    ("foreign_card_2", "Foreign Visitor Card 2 (Title|Desc)", "FF4A235A", "Symbolism of Lord Ganesha|Lord Ganesha is revered as the remover of obstacles and the patron of intellect and wisdom — universal qualities that resonate across cultures."),
    ("foreign_card_3", "Foreign Visitor Card 3 (Title|Desc)", "FF4A235A", "Spiritual Practices|Participate in meditation, chanting of mantras, or attending bhajans (devotional songs) for a firsthand experience of the spiritual atmosphere."),
    ("foreign_card_4", "Foreign Visitor Card 4 (Title|Desc)", "FF4A235A", "Connection with Nature|The temple's natural surroundings — lush greenery, serene atmosphere, and nearby hills — evoke tranquility and oneness with nature."),
    ("foreign_card_5", "Foreign Visitor Card 5 (Title|Desc)", "FF4A235A", "Cultural Exchange|Engage with locals and temple priests to learn customs, traditions, and stories — fostering appreciation for spirituality that transcends borders."),
    ("foreign_card_6", "Foreign Visitor Card 6 (Title|Desc)", "FF4A235A", "Reflection & Contemplation|Take moments of quiet reflection within the temple complex to facilitate personal spiritual insights, regardless of your religious background."),

    # ── SECTION L: Prasad ────────────────────────────────────────────────────
    ("prasad_items",        "Prasad Items (Name|Desc|Name|Desc...)", "FF6D4C41", "Ladoos|Traditional sweet offerings — one of the most popular prasad items at the temple.|Modaks|A beloved sweet associated with Lord Ganesha, considered his favourite offering."),
    ("prasad_cost_note",    "Prasad Cost Note",                      "FF6D4C41", "Varies depending on type and quantity. Check at the temple prasad counter for current prices."),
    ("prasad_purchase_note","Prasad Purchase Note",                  "FF6D4C41", "Prasad can be bought from designated counters within the temple premises."),

    # ── SECTION M: Puja Booking ──────────────────────────────────────────────
    ("booking_online_desc",   "Online Booking Description",   "FF1E8449", "Bookings can be made through the temple's official website. Select from various poojas and rituals available."),
    ("booking_onsite_desc",   "On-Site Booking Description",  "FF1E8449", "Bookings can also be made at the temple office upon arrival. Consult the administration for current rates and availability."),
    ("booking_daily_timing",  "Daily Puja Timing Note",       "FF1E8449", "Morning timings (specific times vary). Contact temple administration for details and current rates."),
    ("booking_festival_note", "Festival Day Booking Note",    "FF1E8449", "Extended hours and special poojas available — timings announced in advance."),

    # ── SECTION N: Sidebar & Administration ─────────────────────────────────
    ("sidebar_nearest_railway","Sidebar: Nearest Railway",    "FF2C3E50", "Khopoli Station (~30 km)"),
    ("sidebar_nearest_airport","Sidebar: Nearest Airport",    "FF2C3E50", "Mumbai Airport (~120 km)"),
    ("sidebar_accessibility",  "Sidebar: Accessibility Note", "FF2C3E50", "Wheelchair accessible|Ramps available|Parking for all vehicles"),
    ("admin_managing_body",    "Admin: Managing Body",        "FF2C3E50", "Ballaleshwar Temple Trust"),
    ("admin_contact",          "Admin: Contact",              "FF2C3E50", "02142-242263"),
    ("admin_address",          "Admin: Full Address",         "FF2C3E50", "Pali Village, Sudhagad Taluka, Raigad District, Maharashtra – 410205"),
]


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "blogs"

    # ── Styles ────
    header_font        = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
    data_font          = Font(name="Calibri", color="FF1A1A1A", size=10)
    example_font       = Font(name="Calibri", color="FF1A5276", size=10, italic=True)
    section_label_font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=9)
    thin_border        = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    wrap_align  = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row 1: Section labels (merged per color group) ───
    # ── Row 2: Column headers ───
    # ── Row 3: Example (Ballaleshwar) data ───
    # ── Row 4+: Empty rows for new blogs ───

    # Build section groups
    sections = []
    current_color = None
    current_start = 1
    section_names = {
        "FF1E3A5F": "🔷  A — Basic Info & SEO",
        "FF2D6A4F": "🔷  B — Hero Section",
        "FF6D4C41": "🔷  C — Quick Stat Cards",
        "FF4A235A": "🔷  D — Overview Tab",
        "FF1A5276": "🔷  E — Facilities Tab",
        "FF117A65": "🔷  F — Guidelines Tab",
        "FF784212": "🔷  G — How to Reach",
        "FF922B21": "🔷  H — YouTube",
        "FF0E6655": "🔷  I — Timings",
        "FF1F618D": "🔷  J — Architecture & Significance",
        "FF4A235A2": "🔷  K — For Foreign Visitors",
        "FF6D4C412": "🔷  L — Prasad",
        "FF1E8449": "🔷  M — Puja Booking",
        "FF2C3E50": "🔷  N — Sidebar & Admin",
    }

    # Track per-column section grouping
    col_meta = []  # (col_idx, key, label, hex_color, example)
    for i, (key, label, color, example) in enumerate(COLUMNS):
        col_meta.append((i + 1, key, label, color, example))

    # ── Write Row 1: Section banners ─────────────────────────────────────────
    prev_color = None
    group_start = 1
    group_label = ""
    # We'll write them after collecting groups
    groups = []

    for i, (col_idx, key, label, color, example) in enumerate(col_meta):
        if color != prev_color:
            if prev_color is not None:
                groups.append((group_start, col_idx - 1, prev_color, group_label))
            # Determine label
            group_label = section_names.get(color, "Section")
            group_start = col_idx
            prev_color = color
    # Last group
    if prev_color is not None:
        groups.append((group_start, len(col_meta), prev_color, group_label))

    # Apply section merged cells + fills on row 1
    for (gs, ge, color, label) in groups:
        start_letter = get_column_letter(gs)
        end_letter   = get_column_letter(ge)
        ws.merge_cells(f"{start_letter}1:{end_letter}1")
        cell = ws[f"{start_letter}1"]
        cell.value = label
        cell.font  = section_label_font
        cell.fill  = PatternFill("solid", fgColor=color)
        cell.alignment = center_align
        cell.border = thin_border

    # ── Write Row 2: Column headers ──────────────────────────────────────────
    for col_idx, key, label, color, example in col_meta:
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f"{key}\n({label})"
        cell.font  = header_font
        cell.fill  = PatternFill("solid", fgColor=color)
        cell.alignment = wrap_align
        cell.border = thin_border

    # ── Write Row 3: Example data (Ballaleshwar) ─────────────────────────────
    for col_idx, key, label, color, example in col_meta:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = example
        cell.font  = example_font
        cell.alignment = wrap_align
        cell.border = thin_border
        # Light blue bg to indicate example
        cell.fill = PatternFill("solid", fgColor="FFE8F4FD")

    # ── Write Row 4-13: Empty template rows ──────────────────────────────────
    for row in range(4, 14):
        for col_idx, key, label, color, example in col_meta:
            cell = ws.cell(row=row, column=col_idx)
            cell.font = data_font
            cell.alignment = wrap_align
            cell.border = thin_border

    # ── Column widths ────────────────────────────────────────────────────────
    long_cols = {
        "meta_description", "keywords", "hero_intro_para_1", "hero_intro_para_2",
        "historical_para_1", "historical_para_2", "spiritual_para_1", "spiritual_para_2",
        "entry_fee_notes", "annual_event_desc", "monthly_event_desc", "donation_channels",
        "facility_1", "facility_2", "facility_3", "facility_4", "facility_5", "facility_6",
        "online_services", "vehicle_pooja_desc", "lost_found_desc",
        "etiquette_points", "photography_rules", "accessibility_rules",
        "air_map_embed", "train_map_embed", "bus_map_embed",
        "arch_style_desc", "arch_material_desc", "arch_scientific_desc", "arch_conservation_desc",
        "significance_para_1", "significance_para_2", "significance_para_3",
        "foreign_card_1", "foreign_card_2", "foreign_card_3",
        "foreign_card_4", "foreign_card_5", "foreign_card_6",
        "prasad_items", "booking_online_desc", "booking_onsite_desc", "admin_address",
        "og_image_url", "canonical_url", "hero_image_url", "address_full",
    }

    for col_idx, key, label, color, example in col_meta:
        col_letter = get_column_letter(col_idx)
        if key in long_cols:
            ws.column_dimensions[col_letter].width = 55
        elif key in {"slug", "date_published", "date_modified"}:
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 28

    # ── Row heights ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 52
    ws.row_dimensions[3].height = 120
    for row in range(4, 14):
        ws.row_dimensions[row].height = 80

    # ── Freeze panes: keep first 3 columns visible while scrolling ───────────
    ws.freeze_panes = "D3"

    # ── Instructions sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("README")
    ws2["A1"] = "📋 BLOG GENERATOR — QUICK START"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=16, color="FF1E3A5F")
    instructions = [
        ("A3",  "STEP 1:", "Install dependencies:  pip install openpyxl jinja2"),
        ("A5",  "STEP 2:", "Fill the 'blogs' sheet — one row per temple (Row 3 is the Ballaleshwar example)"),
        ("A7",  "STEP 3:", "Run:  python scripts/blog_generator/generate_blogs_from_excel.py"),
        ("A9",  "STEP 4:", "Find generated blogs in public/blog/temple/[category]/[slug]/index.html"),
        ("A11", "PIPES:  ", "Use  |  (pipe character) to separate list items in a cell"),
        ("A13", "YOUTUBE:", "Put only the Video ID (e.g. pAtgE0CMKQs), NOT the full URL"),
        ("A15", "MAPS:   ", "Paste only the URL from the Google Maps embed src='' attribute"),
        ("A17", "IMAGES: ", "Use full URL or relative path starting with / or https://"),
        ("A19", "SLUG:   ", "No spaces — use hyphens only (e.g. bhimashankar-jyotirlinga-maharashtra)"),
        ("A21", "ROW 3:  ", "EXAMPLE ROW (Ballaleshwar) — do not delete, use as reference"),
    ]
    for addr, label, text in instructions:
        row_num = int(addr[1:])
        ws2.cell(row=row_num, column=1).value = label
        ws2.cell(row=row_num, column=1).font = Font(name="Calibri", bold=True, size=11, color="FFE23D00")
        ws2.cell(row=row_num, column=2).value = text
        ws2.cell(row=row_num, column=2).font = Font(name="Calibri", size=11)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 80

    wb.save(OUTPUT_PATH)
    print(f"\n[SUCCESS] Excel template created:")
    print(f"    {OUTPUT_PATH}")
    print(f"\n    - Row 1 : Section banners (color-coded)")
    print(f"    - Row 2 : Column headers")
    print(f"    - Row 3 : Ballaleshwar EXAMPLE data (light blue)")
    print(f"    - Row 4+: Empty rows — fill one per temple")
    print(f"\n    Total columns : {len(COLUMNS)}")
    print(f"    Sections      : 14 (A through N)")
    print(f"\n[INFO] Read BLOG_INSTRUCTIONS.md for complete guidance.\n")


if __name__ == "__main__":
    create_template()
